"""Audit the measurement level of the real A-group DPV exports.

The A-group contains three distinct levels that must not be conflated:

* 150 DOE rows in the design workbook;
* 150 CSV time-series files containing instrument summary frames; and
* 150 standard PRT files containing configurable samples of detected
  individual particles.

The audit maps all three by the frozen table/file order, excludes the two
explicit run-063 export-size diagnostics from the run count, and quantifies
how PRT particle summaries relate to CSV frame summaries.  It does not alter
the source workbook or any raw DPV file.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import statistics
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]
DATA_DIR = ROOT / "data-DPV"
WORKBOOK = DATA_DIR / "APS_YSZ_150组总试验文件_A方案_固定观测距离.xlsx"
CSV_DIR = DATA_DIR / "stips-0709"
PRT_DIR = DATA_DIR / "0709"
OUTPUT = (
    HERE
    / "h11_outputs"
    / "dpv_measurement_level"
    / "h11_dpv_measurement_level_audit.json"
)

STANDARD_RUN = re.compile(r"^(?P<run>\d{1,3})\.(?P<suffix>csv|prt)$")
PRT_EXPORT_DIAGNOSTIC = re.compile(
    r"^(?P<run>\d{3})-(?P<sample_size>\d+)\.prt$"
)
CSV_HEADER = (
    "Date/Time(See Manual)",
    "Flow Rate",
    "Temperature",
    "Speed",
    "Diameter",
    "Total Intensity",
    "Peak Intensity",
    "Peak Width",
    "Peak Position",
    "Substrate Temperature",
)
PRT_HEADER = (
    "Date",
    "Time",
    "X",
    "Y",
    "Speed",
    "Temperature",
    "Diameter",
    "Energy A",
    "Energy B",
)


@dataclass(frozen=True)
class DpvMeasurementAuditContract:
    group: str = "A"
    expected_runs: int = 150
    expected_standard_prt_rows_per_run: int = 1000
    fixed_gun_to_dpv_plane_mm: float = 100.0
    execution_order_rule: str = "design_table_order"
    workpiece_present_during_dpv_acquisition: bool = False
    hydrogen_setting: float = 2.5
    hydrogen_reported_unit: str = "g/min"
    carrier_gas_included: bool = False

    def validate(self) -> None:
        if self.group != "A":
            raise ValueError("This audit is restricted to group A")
        if self.expected_runs != 150:
            raise ValueError("The A-group contract must contain 150 runs")
        if self.expected_standard_prt_rows_per_run != 1000:
            raise ValueError("The frozen standard PRT export is 1000 rows")
        if self.fixed_gun_to_dpv_plane_mm != 100.0:
            raise ValueError("The A-group DPV plane must remain at 100 mm")
        if self.execution_order_rule != "design_table_order":
            raise ValueError("A execution order must follow the table order")
        if self.workpiece_present_during_dpv_acquisition:
            raise ValueError("The A-group DPV acquisition had no workpiece")
        if self.hydrogen_setting != 2.5:
            raise ValueError("The reported fixed hydrogen setting changed")
        if self.carrier_gas_included:
            raise ValueError("Unknown carrier gas cannot enter this audit")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _percentile(values: list[float], probability: float) -> float:
    if not values:
        raise ValueError("Cannot calculate a percentile of an empty sample")
    ordered = sorted(values)
    position = probability * (len(ordered) - 1)
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def summarize(values: Iterable[float]) -> dict[str, float | int]:
    finite = [float(value) for value in values if math.isfinite(float(value))]
    if not finite:
        raise ValueError("No finite values were supplied")
    return {
        "n": len(finite),
        "mean": statistics.fmean(finite),
        "standard_deviation": (
            statistics.stdev(finite) if len(finite) > 1 else 0.0
        ),
        "minimum": min(finite),
        "p10": _percentile(finite, 0.10),
        "median": statistics.median(finite),
        "p90": _percentile(finite, 0.90),
        "maximum": max(finite),
    }


def pearson(x: list[float], y: list[float]) -> float:
    if len(x) != len(y) or len(x) < 2:
        raise ValueError("Pearson inputs must have equal length >= 2")
    x_mean = statistics.fmean(x)
    y_mean = statistics.fmean(y)
    numerator = sum(
        (left - x_mean) * (right - y_mean)
        for left, right in zip(x, y)
    )
    denominator = math.sqrt(
        sum((value - x_mean) ** 2 for value in x)
        * sum((value - y_mean) ** 2 for value in y)
    )
    if denominator == 0:
        raise ValueError("Pearson correlation is undefined for a constant")
    return numerator / denominator


def _average_ranks(values: list[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    start = 0
    while start < len(indexed):
        end = start + 1
        while end < len(indexed) and indexed[end][1] == indexed[start][1]:
            end += 1
        rank = (start + 1 + end) / 2
        for position in range(start, end):
            ranks[indexed[position][0]] = rank
        start = end
    return ranks


def spearman(x: list[float], y: list[float]) -> float:
    return pearson(_average_ranks(x), _average_ranks(y))


def read_design(path: Path, contract: DpvMeasurementAuditContract) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["150组总试验表_A方案"]
        values = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    header = list(values[3])
    expected = (
        "序号",
        "试验编号",
        "实际运行顺序",
        "DOE模块",
        "试验目的",
        "电流(A)",
        "氩气流量(scfh)",
        "喷涂距离d(mm)",
        "送粉速率设定(g/min)",
    )
    if tuple(header[: len(expected)]) != expected:
        raise ValueError("Unexpected A-group design header")
    records: list[dict[str, Any]] = []
    for row in values[4:]:
        if row[0] is None:
            continue
        run = int(row[0])
        records.append(
            {
                "run": run,
                "experiment_id": str(row[1]),
                "recorded_execution_order": row[2],
                "effective_execution_order": run,
                "doe_module": str(row[3]).split("｜")[0],
                "purpose": str(row[4]),
                "current_a": float(row[5]),
                "argon_scfh": float(row[6]),
                "spray_distance_mm": float(row[7]),
                "powder_feed_g_min": float(row[8]),
                "measurement_scheme": str(row[10]),
                "observation_mode": str(row[11]),
                "workbook_observation_distance_mm": row[12],
                "data_quality": str(row[21]),
                "hydrogen_setting": contract.hydrogen_setting,
                "hydrogen_reported_unit": contract.hydrogen_reported_unit,
                "hydrogen_to_argon_reported_setting_ratio": (
                    contract.hydrogen_setting / float(row[6])
                ),
            }
        )
    return records


def _standard_files(directory: Path, suffix: str) -> dict[int, Path]:
    files: dict[int, Path] = {}
    for path in directory.iterdir():
        match = STANDARD_RUN.match(path.name)
        if match is None or match.group("suffix") != suffix:
            continue
        run = int(match.group("run"))
        if run in files:
            raise ValueError(f"Duplicate standard {suffix} run {run}")
        files[run] = path
    return files


def read_prt(path: Path) -> dict[str, Any]:
    lines = path.read_text(encoding="utf-8", errors="strict").splitlines()
    if not lines:
        raise ValueError(f"Empty PRT file: {path}")
    normalized_header = tuple(lines[0].split())
    if normalized_header != (
        "Date",
        "Time",
        "X",
        "Y",
        "Speed",
        "Temperature",
        "Diameter",
        "Energy",
        "A",
        "Energy",
        "B",
    ):
        raise ValueError(f"Unexpected PRT header: {path}")
    speed: list[float] = []
    temperature: list[float] = []
    diameter: list[float] = []
    x_position: list[float] = []
    y_position: list[float] = []
    energy_a: list[float] = []
    energy_b: list[float] = []
    timestamps: list[datetime] = []
    for line_number, line in enumerate(lines[1:], start=2):
        fields = line.split()
        if len(fields) != 9:
            raise ValueError(
                f"{path}:{line_number} has {len(fields)} fields, expected 9"
            )
        timestamps.append(
            datetime.strptime(
                f"{fields[0]} {fields[1]}",
                "%m/%d/%Y %H:%M:%S:%f",
            )
        )
        x_position.append(float(fields[2]))
        y_position.append(float(fields[3]))
        speed.append(float(fields[4]))
        temperature.append(float(fields[5]))
        diameter.append(float(fields[6]))
        energy_a.append(float(fields[7]))
        energy_b.append(float(fields[8]))
    if not speed:
        raise ValueError(f"PRT file contains no particles: {path}")
    return {
        "rows": len(speed),
        "unique_timestamps": len(set(timestamps)),
        "duration_s": (
            (max(timestamps) - min(timestamps)).total_seconds()
            if len(timestamps) > 1
            else 0.0
        ),
        "speed_m_s": summarize(speed),
        "temperature_c": summarize(temperature),
        "diameter_um": summarize(diameter),
        "x_position": summarize(x_position),
        "y_position": summarize(y_position),
        "energy_a": summarize(energy_a),
        "energy_b": summarize(energy_b),
        "nonpositive_counts": {
            "speed": sum(value <= 0 for value in speed),
            "temperature": sum(value <= 0 for value in temperature),
            "diameter": sum(value <= 0 for value in diameter),
        },
    }


def read_csv_frames(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if tuple(reader.fieldnames or ()) != CSV_HEADER:
            raise ValueError(f"Unexpected CSV header: {path}")
        rows = list(reader)
    if not rows:
        raise ValueError(f"CSV file contains no frames: {path}")
    time_serial = [float(row[CSV_HEADER[0]]) for row in rows]
    valid = [
        row
        for row in rows
        if float(row["Temperature"]) > 0
        and float(row["Speed"]) > 0
        and float(row["Diameter"]) > 0
    ]
    if not valid:
        raise ValueError(f"CSV file contains no valid DPV frames: {path}")
    cadence = [
        (right - left) * 86400.0
        for left, right in zip(time_serial, time_serial[1:])
        if right > left
    ]
    intensity_columns = (
        "Total Intensity",
        "Peak Intensity",
        "Peak Width",
        "Peak Position",
        "Substrate Temperature",
    )
    return {
        "rows": len(rows),
        "valid_rows": len(valid),
        "zero_primary_rows": len(rows) - len(valid),
        "duration_s": (max(time_serial) - min(time_serial)) * 86400.0,
        "median_cadence_s": statistics.median(cadence) if cadence else 0.0,
        "flow_rate": summarize(float(row["Flow Rate"]) for row in rows),
        "temperature_c": summarize(
            float(row["Temperature"]) for row in valid
        ),
        "speed_m_s": summarize(float(row["Speed"]) for row in valid),
        "diameter_um": summarize(float(row["Diameter"]) for row in valid),
        "all_auxiliary_measurement_columns_zero": all(
            float(row[column]) == 0
            for row in rows
            for column in intensity_columns
        ),
    }


def _comparison(
    particle_runs: list[dict[str, Any]],
    frame_runs: list[dict[str, Any]],
    quantity: str,
) -> dict[str, Any]:
    particle = [
        float(run[quantity]["mean"]) for run in particle_runs
    ]
    frame = [float(run[quantity]["mean"]) for run in frame_runs]
    relative = [
        abs(right - left) / max(abs(left), 1e-12)
        for left, right in zip(particle, frame)
    ]
    return {
        "n_runs": len(particle),
        "pearson": pearson(particle, frame),
        "spearman": spearman(particle, frame),
        "absolute_relative_difference": summarize(relative),
    }


def build_audit(
    *,
    contract: DpvMeasurementAuditContract,
    design: list[dict[str, Any]],
    standard_prt: dict[int, Path],
    standard_csv: dict[int, Path],
    prt_diagnostics: list[dict[str, Any]],
) -> dict[str, Any]:
    expected_runs = list(range(1, contract.expected_runs + 1))
    particle_runs = [read_prt(standard_prt[run]) for run in expected_runs]
    frame_runs = [read_csv_frames(standard_csv[run]) for run in expected_runs]
    settings = {
        (
            row["current_a"],
            row["argon_scfh"],
            row["spray_distance_mm"],
            row["powder_feed_g_min"],
        )
        for row in design
    }
    gates = {
        "design_has_exactly_150_rows": len(design) == contract.expected_runs,
        "design_sequence_is_1_to_150": (
            [row["run"] for row in design] == expected_runs
        ),
        "all_measurement_scheme_A": all(
            row["measurement_scheme"] == "A" for row in design
        ),
        "all_observation_modes_fixed": all(
            row["observation_mode"] == "固定观测距离" for row in design
        ),
        "workbook_observation_distance_blank_for_all_runs": all(
            row["workbook_observation_distance_mm"] is None for row in design
        ),
        "all_design_quality_marked_valid": all(
            row["data_quality"] == "有效" for row in design
        ),
        "standard_prt_files_exactly_001_to_150": (
            sorted(standard_prt) == expected_runs
        ),
        "standard_csv_files_exactly_001_to_150": (
            sorted(standard_csv) == expected_runs
        ),
        "all_standard_prt_exports_have_1000_particles": all(
            run["rows"] == contract.expected_standard_prt_rows_per_run
            for run in particle_runs
        ),
        "all_prt_primary_particle_values_positive": all(
            all(count == 0 for count in run["nonpositive_counts"].values())
            for run in particle_runs
        ),
        "every_csv_has_at_least_one_valid_frame": all(
            run["valid_rows"] > 0 for run in frame_runs
        ),
        "fixed_dpv_plane_defined_by_user_contract": (
            contract.fixed_gun_to_dpv_plane_mm == 100.0
        ),
        "workpiece_absent_in_measurement_branch": (
            not contract.workpiece_present_during_dpv_acquisition
        ),
    }
    return {
        "schema_version": "h11_dpv_measurement_level_audit_v1",
        "status": (
            "pass_A_group_measurement_level_audit"
            if all(gates.values())
            else "fail_A_group_measurement_level_audit"
        ),
        "contract": asdict(contract),
        "gates": gates,
        "design": {
            "runs": len(design),
            "unique_four_factor_settings": len(settings),
            "module_counts": dict(
                sorted(Counter(row["doe_module"] for row in design).items())
            ),
            "current_a": summarize(row["current_a"] for row in design),
            "argon_scfh": summarize(row["argon_scfh"] for row in design),
            "spray_distance_mm": summarize(
                row["spray_distance_mm"] for row in design
            ),
            "powder_feed_g_min": summarize(
                row["powder_feed_g_min"] for row in design
            ),
            "hydrogen_to_argon_reported_setting_ratio": summarize(
                row["hydrogen_to_argon_reported_setting_ratio"]
                for row in design
            ),
            "ratio_interpretation": (
                "Reported-setting ratio only. Hydrogen is reported as g/min "
                "and argon as scfh, so this is not a physical composition "
                "ratio until units and gas conditions are reconciled."
            ),
        },
        "particle_exports": {
            "standard_run_files": len(standard_prt),
            "standard_rows_per_run": sorted(
                {run["rows"] for run in particle_runs}
            ),
            "total_standard_particle_records": sum(
                run["rows"] for run in particle_runs
            ),
            "per_run_unique_timestamp": summarize(
                run["unique_timestamps"] for run in particle_runs
            ),
            "per_run_duration_s": summarize(
                run["duration_s"] for run in particle_runs
            ),
            "run_mean_temperature_c": summarize(
                run["temperature_c"]["mean"] for run in particle_runs
            ),
            "run_mean_speed_m_s": summarize(
                run["speed_m_s"]["mean"] for run in particle_runs
            ),
            "run_mean_diameter_um": summarize(
                run["diameter_um"]["mean"] for run in particle_runs
            ),
            "export_size_diagnostics": prt_diagnostics,
            "interpretation": (
                "Each standard PRT contains individual detected-particle "
                "records. The 10/1000/10000 exports for run 063 demonstrate "
                "a configurable retained sample size; 1000 is not the total "
                "number of particles crossing the measurement plane."
            ),
        },
        "frame_exports": {
            "run_files": len(standard_csv),
            "total_frames": sum(run["rows"] for run in frame_runs),
            "total_valid_frames": sum(
                run["valid_rows"] for run in frame_runs
            ),
            "per_run_frames": summarize(run["rows"] for run in frame_runs),
            "per_run_valid_frames": summarize(
                run["valid_rows"] for run in frame_runs
            ),
            "median_cadence_s": summarize(
                run["median_cadence_s"] for run in frame_runs
            ),
            "run_mean_temperature_c": summarize(
                run["temperature_c"]["mean"] for run in frame_runs
            ),
            "run_mean_speed_m_s": summarize(
                run["speed_m_s"]["mean"] for run in frame_runs
            ),
            "run_mean_diameter_um": summarize(
                run["diameter_um"]["mean"] for run in frame_runs
            ),
            "runs_with_all_auxiliary_columns_zero": sum(
                run["all_auxiliary_measurement_columns_zero"]
                for run in frame_runs
            ),
            "interpretation": (
                "CSV rows are time-indexed instrument summary frames, not "
                "individual particles."
            ),
        },
        "particle_vs_frame_run_mean": {
            quantity: _comparison(particle_runs, frame_runs, quantity)
            for quantity in (
                "temperature_c",
                "speed_m_s",
                "diameter_um",
            )
        },
        "observation_operator_contract": {
            "stage_1": (
                "Simulate a 7YSZ particle population crossing the fixed "
                "gun-relative z=100 mm plane."
            ),
            "stage_2": (
                "Apply a detection/retention operator in particle position, "
                "diameter, temperature, speed, and signal space."
            ),
            "stage_3": (
                "Compare detected-particle distribution summaries with PRT; "
                "then aggregate the retained population to CSV-frame or "
                "run-level statistics."
            ),
            "prt_is_individual_particle_sample": True,
            "csv_is_summary_frame_series": True,
            "raw_source_files_modified": False,
            "exact_instrument_thresholds_known": False,
            "paper_prediction_allowed": False,
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--csv-dir", type=Path, default=CSV_DIR)
    parser.add_argument("--prt-dir", type=Path, default=PRT_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT)
    args = parser.parse_args()
    for path in (args.workbook, args.csv_dir, args.prt_dir):
        if not path.exists():
            raise FileNotFoundError(path)

    contract = DpvMeasurementAuditContract()
    contract.validate()
    design = read_design(args.workbook, contract)
    standard_prt = _standard_files(args.prt_dir, "prt")
    standard_csv = _standard_files(args.csv_dir, "csv")
    diagnostics: list[dict[str, Any]] = []
    for path in sorted(args.prt_dir.iterdir()):
        match = PRT_EXPORT_DIAGNOSTIC.match(path.name)
        if match is None:
            continue
        metrics = read_prt(path)
        diagnostics.append(
            {
                "file": path.name,
                "run": int(match.group("run")),
                "declared_sample_size": int(match.group("sample_size")),
                "actual_rows": metrics["rows"],
                "sha256": sha256(path),
            }
        )
    audit = build_audit(
        contract=contract,
        design=design,
        standard_prt=standard_prt,
        standard_csv=standard_csv,
        prt_diagnostics=diagnostics,
    )
    audit["sources"] = {
        "workbook": str(args.workbook.resolve()),
        "workbook_sha256": sha256(args.workbook),
        "csv_directory": str(args.csv_dir.resolve()),
        "prt_directory": str(args.prt_dir.resolve()),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(audit, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Wrote DPV measurement-level audit: {args.output}")
    print(f"Status: {audit['status']}")
    return int(audit["status"].startswith("fail_"))


if __name__ == "__main__":
    raise SystemExit(main())
